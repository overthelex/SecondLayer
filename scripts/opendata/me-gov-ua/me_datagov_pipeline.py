#!/usr/bin/env python3
"""
me.gov.ua (Ministry of Economy) open-data mirror pipeline.

Mirrors ALL open datasets published by the Ministry of Economy of Ukraine on
data.gov.ua into the openreyestr Postgres DB (tables me_datasets / me_resources /
me_records — see mcp_openreyestr/src/migrations/015_me_gov_datasets_mirror.sql).

CKAN org: ministerstvo-ekonomichnoho-rozvytku-i-torhivli-ukrayiny
          id c9edbe3d-0a05-4d00-b23b-7c1fe5810ca4  (69 datasets)

Stages (each is idempotent / resumable):
    discover  CKAN package_search -> upsert me_datasets + me_resources
    download  fetch resource files to STAGING_DIR, sha256, mark 'downloaded'
    import    parse tabular files (CSV/XLSX/JSON/XML) -> me_records, mark 'imported'
              non-tabular (PDF/DOC/ZIP/...) -> mark 'skipped' (file mirrored only)
    verify    counts + reconciliation report (also runs a prod-side sanity query)
    all       discover -> download -> import -> verify

Usage:
    # local DB (default 5435)
    python3 me_datagov_pipeline.py all
    python3 me_datagov_pipeline.py discover
    python3 me_datagov_pipeline.py download --limit 50
    python3 me_datagov_pipeline.py import
    python3 me_datagov_pipeline.py verify

    # prod DB over WG tunnel (openreyestr_prod on 127.0.0.1:5440)
    POSTGRES_HOST=127.0.0.1 POSTGRES_PORT=5440 POSTGRES_DB=openreyestr_prod \\
    POSTGRES_USER=openreyestr POSTGRES_PASSWORD=... \\
    python3 me_datagov_pipeline.py verify

Env (same names as mcp_openreyestr):
    POSTGRES_HOST (localhost)  POSTGRES_PORT (5435)  POSTGRES_DB (openreyestr)
    POSTGRES_USER (openreyestr)  POSTGRES_PASSWORD (openreyestr_password)
    DATABASE_URL  (overrides the above if set)
    ME_STAGING_DIR (default /data/opendata/me_gov_ua on prod host, else ./staging)

Deps: pip install -r requirements.txt
      (psycopg2-binary requests aiohttp aiofiles openpyxl)
"""

import argparse
import asyncio
import hashlib
import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests

try:
    import aiohttp
    import aiofiles
except ImportError:
    aiohttp = None  # only needed for the download stage

import psycopg2
import psycopg2.extras

# ── Config ──────────────────────────────────────────────────────────────────
CKAN_API = "https://data.gov.ua/api/3/action"
ORG_SLUG = "ministerstvo-ekonomichnoho-rozvytku-i-torhivli-ukrayiny"
ORG_ID = "c9edbe3d-0a05-4d00-b23b-7c1fe5810ca4"
USER_AGENT = "SecondLayer-Legal-Platform/1.0 (+https://legal.org.ua)"

# Multi-IP source binding (AWS ENIs on the local data-proxy box). Falls back to
# the default outbound IP when none of these are bound on the host.
SOURCE_IPS = [
    "172.31.29.20",
    "172.31.21.255",
    "172.31.31.40",
    "172.31.22.206",
    "172.31.28.109",
]
THREADS_PER_IP = 5

STAGING_DIR = Path(
    os.environ.get(
        "ME_STAGING_DIR",
        "/data/opendata/me_gov_ua"
        if Path("/data/opendata").is_dir()
        else "./staging",
    )
)

# Formats we parse into me_records. Everything else is mirrored as a file only.
TABULAR_FORMATS = {"CSV", "TSV", "XLSX", "XLS", "JSON", "XML"}
ARCHIVE_FORMATS = {"7Z", "ZIP"}  # mirrored as file; extraction is a later pass
MAX_ROWS_PER_RESOURCE = int(os.environ.get("ME_MAX_ROWS", "500000"))  # safety cap

# data.gov.ua format labels are hand-entered and frequently use Cyrillic
# homoglyphs (Х/Р/С/… instead of X/P/C/…), e.g. "ХLSX", "РDF", "СSV". Normalize
# them so tabular resources aren't misclassified and skipped.
_CYR2LAT = str.maketrans(
    {"Х": "X", "Р": "P", "С": "C", "А": "A", "В": "B", "Е": "E", "К": "K",
     "М": "M", "О": "O", "Т": "T", "Н": "H", "І": "I", "Ѕ": "S", "Ј": "J",
     "х": "X", "р": "P", "с": "C", "а": "A", "в": "B", "е": "E", "к": "K",
     "м": "M", "о": "O", "т": "T", "н": "H", "і": "I"}
)
_EXT2FMT = {
    ".xlsx": "XLSX", ".xls": "XLS", ".csv": "CSV", ".tsv": "TSV",
    ".json": "JSON", ".xml": "XML", ".pdf": "PDF", ".doc": "DOC",
    ".docx": "DOCX", ".rtf": "RTF", ".7z": "7Z", ".zip": "ZIP",
}


def _norm_fmt(s: str | None) -> str:
    return (s or "").translate(_CYR2LAT).upper().strip()


def detect_format(local_path: str | None, ckan_format: str | None) -> str:
    """Determine the true resource format from the file extension first
    (robust against Cyrillic-homoglyph CKAN labels), then the CKAN label.
    For composite labels like '7Z,XLSX' the archive wins (the file on disk is
    the .7z), so it is mirrored rather than mis-parsed."""
    if local_path:
        ext = Path(local_path).suffix.lower()
        # a Cyrillic-tainted extension won't match; fall through to the label
        if ext in _EXT2FMT:
            return _EXT2FMT[ext]
    norm = _norm_fmt(ckan_format)
    tokens = [t for t in re.split(r"[,\s]+", norm) if t]
    for t in tokens:
        if t in ARCHIVE_FORMATS:
            return t
    return tokens[0] if tokens else ""

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("me_datagov")


# ── DB helpers ──────────────────────────────────────────────────────────────
def connect():
    dsn = os.environ.get("DATABASE_URL")
    if dsn:
        conn = psycopg2.connect(dsn)
    else:
        conn = psycopg2.connect(
            host=os.environ.get("POSTGRES_HOST", "localhost"),
            port=int(os.environ.get("POSTGRES_PORT", "5435")),
            dbname=os.environ.get("POSTGRES_DB", "openreyestr"),
            user=os.environ.get("POSTGRES_USER", "openreyestr"),
            password=os.environ.get("POSTGRES_PASSWORD", "openreyestr_password"),
        )
    conn.autocommit = False
    return conn


def _ts(value):
    """Parse a CKAN ISO timestamp to a tz-aware datetime, or None."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except (ValueError, AttributeError):
        return None


# ── discover ────────────────────────────────────────────────────────────────
def ckan_get(action: str, **params) -> dict:
    r = requests.get(
        f"{CKAN_API}/{action}",
        params=params,
        headers={"User-Agent": USER_AGENT},
        timeout=60,
    )
    r.raise_for_status()
    data = r.json()
    if not data.get("success"):
        raise RuntimeError(f"CKAN {action} failed: {data}")
    return data["result"]


def discover(conn):
    """Enumerate every package of the org and upsert metadata + resources."""
    log.info("Discovering datasets for org %s ...", ORG_SLUG)
    packages: list[dict] = []
    start, rows = 0, 100
    while True:
        result = ckan_get(
            "package_search",
            fq=f"organization:{ORG_SLUG}",
            rows=rows,
            start=start,
        )
        batch = result.get("results", [])
        packages.extend(batch)
        total = result.get("count", 0)
        log.info("  fetched %d/%d packages", len(packages), total)
        start += rows
        if start >= total or not batch:
            break

    ds_upserts = res_upserts = 0
    with conn.cursor() as cur:
        for pkg in packages:
            # package_search already returns resources, but fetch package_show to be
            # authoritative (some CKAN instances trim resources in search results).
            full = pkg
            if not full.get("resources"):
                try:
                    full = ckan_get("package_show", id=pkg["id"])
                except Exception as e:  # noqa: BLE001
                    log.warning("  package_show failed for %s: %s", pkg.get("name"), e)

            tags = [t.get("name") for t in full.get("tags", []) if t.get("name")]
            cur.execute(
                """
                INSERT INTO me_datasets
                    (ckan_id, slug, title, notes, org_id, tags, num_resources,
                     ckan_metadata_created, ckan_metadata_modified, raw, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW())
                ON CONFLICT (ckan_id) DO UPDATE SET
                    slug = EXCLUDED.slug,
                    title = EXCLUDED.title,
                    notes = EXCLUDED.notes,
                    org_id = EXCLUDED.org_id,
                    tags = EXCLUDED.tags,
                    num_resources = EXCLUDED.num_resources,
                    ckan_metadata_created = EXCLUDED.ckan_metadata_created,
                    ckan_metadata_modified = EXCLUDED.ckan_metadata_modified,
                    raw = EXCLUDED.raw,
                    updated_at = NOW()
                RETURNING id
                """,
                (
                    full["id"],
                    full.get("name"),
                    full.get("title") or full.get("name") or full["id"],
                    full.get("notes"),
                    ORG_ID,
                    tags,
                    len(full.get("resources", [])),
                    _ts(full.get("metadata_created")),
                    _ts(full.get("metadata_modified")),
                    json.dumps(full, ensure_ascii=False),
                ),
            )
            dataset_id = cur.fetchone()[0]
            ds_upserts += 1

            for res in full.get("resources", []):
                if not res.get("url"):
                    continue
                cur.execute(
                    """
                    INSERT INTO me_resources
                        (dataset_id, ckan_resource_id, name, format, url, mimetype,
                         size_bytes, ckan_created, ckan_last_modified, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW())
                    ON CONFLICT (ckan_resource_id) DO UPDATE SET
                        dataset_id = EXCLUDED.dataset_id,
                        name = EXCLUDED.name,
                        format = EXCLUDED.format,
                        url = EXCLUDED.url,
                        mimetype = EXCLUDED.mimetype,
                        size_bytes = EXCLUDED.size_bytes,
                        ckan_created = EXCLUDED.ckan_created,
                        ckan_last_modified = EXCLUDED.ckan_last_modified,
                        updated_at = NOW()
                    """,
                    (
                        dataset_id,
                        res["id"],
                        res.get("name"),
                        (res.get("format") or "").upper() or None,
                        res["url"],
                        res.get("mimetype"),
                        int(res["size"]) if str(res.get("size") or "").isdigit() else None,
                        _ts(res.get("created")),
                        _ts(res.get("last_modified")) or _ts(res.get("metadata_modified")),
                    ),
                )
                res_upserts += 1
    conn.commit()
    log.info("discover done: %d datasets, %d resources upserted", ds_upserts, res_upserts)


# ── download ────────────────────────────────────────────────────────────────
def _available_ips() -> list[str]:
    import subprocess

    try:
        out = subprocess.run(
            ["ip", "-4", "addr", "show"], capture_output=True, text=True
        ).stdout
    except FileNotFoundError:
        return ["0.0.0.0"]
    ips = [
        ln.strip().split()[1].split("/")[0]
        for ln in out.splitlines()
        if ln.strip().startswith("inet ")
    ]
    bound = [ip for ip in SOURCE_IPS if ip in ips]
    return bound or ["0.0.0.0"]


def _resource_filename(res_id: str, url: str, fmt: str | None) -> str:
    tail = url.split("?")[0].split("/")[-1]
    if not tail or tail == "download":
        ext = (_norm_fmt(fmt).split(",")[0].strip() or "BIN").lower()
        tail = f"{res_id}.{ext}"
    return f"{res_id}__{tail}"  # prefix with resource id to guarantee uniqueness


def download(conn, limit: int | None):
    if aiohttp is None:
        log.error("aiohttp/aiofiles not installed. pip install aiohttp aiofiles")
        sys.exit(1)

    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            """
            SELECT id, ckan_resource_id, url, format, local_path
            FROM me_resources
            WHERE import_status IN ('pending', 'error')
               OR local_path IS NULL
            ORDER BY id
            %s
            """
            % (f"LIMIT {int(limit)}" if limit else "")
        )
        pending = cur.fetchall()

    if not pending:
        log.info("download: nothing pending")
        return

    ips = _available_ips()
    log.info(
        "download: %d resources, %d source IP(s) × %d threads",
        len(pending), len(ips), THREADS_PER_IP,
    )

    results: dict[int, dict] = {}

    async def fetch_one(session, res, sem):
        fname = _resource_filename(res["ckan_resource_id"], res["url"], res["format"])
        outpath = STAGING_DIR / fname
        if outpath.exists() and outpath.stat().st_size > 0:
            sha = _sha256(outpath)
            results[res["id"]] = {"local_path": str(outpath), "sha256": sha,
                                  "status": "downloaded"}
            return
        async with sem:
            for attempt in range(3):
                try:
                    to = aiohttp.ClientTimeout(total=600)
                    async with session.get(res["url"], timeout=to,
                                           allow_redirects=True) as r:
                        if r.status == 200:
                            data = await r.read()
                            if data:
                                async with aiofiles.open(outpath, "wb") as f:
                                    await f.write(data)
                                results[res["id"]] = {
                                    "local_path": str(outpath),
                                    "sha256": hashlib.sha256(data).hexdigest(),
                                    "status": "downloaded",
                                }
                                return
                        elif r.status == 429:
                            await asyncio.sleep(10 * (attempt + 1))
                            continue
                        elif r.status >= 500:
                            await asyncio.sleep(5 * (attempt + 1))
                            continue
                        else:
                            log.warning("HTTP %s for %s", r.status, fname)
                            break
                except Exception as e:  # noqa: BLE001
                    log.warning("download error %s: %s (try %d)", fname, e, attempt + 1)
                    await asyncio.sleep(3 * (attempt + 1))
            results[res["id"]] = {"status": "error",
                                  "error": "download failed after 3 attempts"}

    async def run():
        # round-robin resources across IPs
        buckets: dict[str, list] = {ip: [] for ip in ips}
        for i, res in enumerate(pending):
            buckets[ips[i % len(ips)]].append(res)

        async def ip_worker(ip, items):
            sem = asyncio.Semaphore(THREADS_PER_IP)
            conn_kw = {}
            if ip != "0.0.0.0":
                conn_kw["local_addr"] = (ip, 0)
            connector = aiohttp.TCPConnector(limit=THREADS_PER_IP, **conn_kw)
            async with aiohttp.ClientSession(
                connector=connector, headers={"User-Agent": USER_AGENT}
            ) as session:
                await asyncio.gather(*(fetch_one(session, r, sem) for r in items))

        await asyncio.gather(*(ip_worker(ip, it) for ip, it in buckets.items()))

    asyncio.run(run())

    ok = err = 0
    with conn.cursor() as cur:
        for rid, r in results.items():
            if r["status"] == "downloaded":
                cur.execute(
                    """UPDATE me_resources SET local_path=%s, sha256=%s,
                       downloaded_at=NOW(), import_status='downloaded',
                       import_error=NULL, updated_at=NOW() WHERE id=%s""",
                    (r["local_path"], r["sha256"], rid),
                )
                ok += 1
            else:
                cur.execute(
                    """UPDATE me_resources SET import_status='error',
                       import_error=%s, updated_at=NOW() WHERE id=%s""",
                    (r.get("error"), rid),
                )
                err += 1
    conn.commit()
    log.info("download done: %d ok, %d error", ok, err)


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ── import (parse tabular -> me_records) ────────────────────────────────────
def _read_text(path: Path) -> str:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "cp1251", "windows-1251"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _rows_from_csv(path: Path) -> list[dict]:
    import csv

    text = _read_text(path)
    lines = text.splitlines()
    header = lines[0] if lines else ""
    # UA gov CSVs are overwhelmingly ';'-delimited (',' is the decimal
    # separator), which fools csv.Sniffer. Pick the delimiter that occurs most
    # in the header row; break ties in favour of ';'.
    delim = max((";", "\t", "|", ","), key=lambda d: (header.count(d), d == ";"))
    if header.count(delim) == 0:
        delim = ","
    reader = csv.DictReader(lines, delimiter=delim)
    return [dict(row) for row in reader]


def _rows_from_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            continue
        cols = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
        for r in it:
            rows.append(
                {cols[i]: (v.isoformat() if hasattr(v, "isoformat") else v)
                 for i, v in enumerate(r) if i < len(cols)}
            )
            if len(rows) >= MAX_ROWS_PER_RESOURCE:
                return rows
    return rows


def _rows_from_json(path: Path) -> list[dict]:
    data = json.loads(_read_text(path))
    if isinstance(data, list):
        return [d if isinstance(d, dict) else {"value": d} for d in data]
    if isinstance(data, dict):
        for key in ("data", "records", "result", "items", "rows"):
            if isinstance(data.get(key), list):
                return [d if isinstance(d, dict) else {"value": d} for d in data[key]]
        return [data]
    return [{"value": data}]


def _rows_from_xml(path: Path) -> list[dict]:
    import xml.etree.ElementTree as ET

    def strip_ns(tag: str) -> str:
        return tag.split("}", 1)[-1] if "}" in tag else tag

    def elem_to_dict(el):
        d = {strip_ns(k): v for k, v in el.attrib.items()}
        for child in el:
            key = strip_ns(child.tag)
            val = elem_to_dict(child) if len(child) else (child.text or "").strip()
            if key in d:
                if not isinstance(d[key], list):
                    d[key] = [d[key]]
                d[key].append(val)
            else:
                d[key] = val
        if not d and (el.text or "").strip():
            return (el.text or "").strip()
        return d

    root = ET.parse(path).getroot()
    children = list(root)
    # Heuristic: repeating same-tag children = record rows.
    if children:
        tags = {strip_ns(c.tag) for c in children}
        if len(tags) == 1 and len(children) > 1:
            return [elem_to_dict(c) if isinstance(elem_to_dict(c), dict)
                    else {"value": elem_to_dict(c)} for c in children]
    top = elem_to_dict(root)
    return [top if isinstance(top, dict) else {"value": top}]


PARSERS = {
    "CSV": _rows_from_csv,
    "TSV": _rows_from_csv,
    "XLSX": _rows_from_xlsx,
    "XLS": _rows_from_xlsx,
    "JSON": _rows_from_json,
    "XML": _rows_from_xml,
}


def import_records(conn, reimport: bool):
    statuses = "('downloaded', 'imported', 'error')" if reimport else "('downloaded')"
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            f"""
            SELECT id, ckan_resource_id, format, local_path
            FROM me_resources
            WHERE local_path IS NOT NULL AND import_status IN {statuses}
            ORDER BY id
            """
        )
        todo = cur.fetchall()

    log.info("import: %d resources to process", len(todo))
    imported = skipped = errored = 0

    for res in todo:
        path = Path(res["local_path"])
        if not path.exists():
            _mark(conn, res["id"], "error", err="local file missing")
            errored += 1
            continue
        fmt = detect_format(res["local_path"], res["format"])
        if fmt not in TABULAR_FORMATS:
            # non-tabular (PDF/DOC/RTF) or archive (7Z/ZIP) — mirrored as file only
            _mark(conn, res["id"], "skipped", rows=0)
            skipped += 1
            continue
        try:
            rows = PARSERS[fmt](path)
        except Exception as e:  # noqa: BLE001
            log.warning("parse failed id=%s (%s): %s", res["id"], fmt, e)
            _mark(conn, res["id"], "error", err=f"parse: {e}")
            errored += 1
            continue

        rows = rows[:MAX_ROWS_PER_RESOURCE]
        with conn.cursor() as cur:
            cur.execute("DELETE FROM me_records WHERE resource_id=%s", (res["id"],))
            if rows:
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO me_records (resource_id, row_index, data)
                       VALUES %s
                       ON CONFLICT (resource_id, row_index) DO UPDATE
                       SET data = EXCLUDED.data""",
                    [
                        (res["id"], i, json.dumps(_jsonable(row), ensure_ascii=False))
                        for i, row in enumerate(rows)
                    ],
                    template="(%s, %s, %s::jsonb)",
                    page_size=1000,
                )
            cur.execute(
                """UPDATE me_resources SET import_status='imported', row_count=%s,
                   import_error=NULL, updated_at=NOW() WHERE id=%s""",
                (len(rows), res["id"]),
            )
        conn.commit()
        imported += 1
        log.info("  imported id=%s %s rows=%d", res["id"], fmt, len(rows))

    log.info("import done: %d imported, %d skipped(non-tabular), %d error",
             imported, skipped, errored)


def _jsonable(row: dict) -> dict:
    out = {}
    for k, v in row.items():
        key = str(k).strip() if k is not None else "col"
        if hasattr(v, "isoformat"):
            v = v.isoformat()
        elif isinstance(v, bytes):
            v = v.decode("utf-8", "replace")
        out[key] = v
    return out


def _mark(conn, res_id: int, status: str, rows: int | None = None, err=None):
    with conn.cursor() as cur:
        cur.execute(
            """UPDATE me_resources SET import_status=%s, row_count=%s,
               import_error=%s, updated_at=NOW() WHERE id=%s""",
            (status, rows, err, res_id),
        )
    conn.commit()


# ── verify ──────────────────────────────────────────────────────────────────
def verify(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT count(*) FROM me_datasets")
        n_datasets = cur.fetchone()[0]
        cur.execute("SELECT count(*), coalesce(sum(num_resources),0) FROM me_datasets")
        _, expected_res = cur.fetchone()
        cur.execute("SELECT count(*) FROM me_resources")
        n_res = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM me_records")
        n_records = cur.fetchone()[0]

        cur.execute(
            "SELECT import_status, count(*) FROM me_resources GROUP BY 1 ORDER BY 1"
        )
        by_status = cur.fetchall()
        cur.execute(
            """SELECT coalesce(format,'(none)'), count(*)
               FROM me_resources GROUP BY 1 ORDER BY 2 DESC"""
        )
        by_format = cur.fetchall()
        cur.execute(
            """SELECT d.title, count(r.*) AS res, coalesce(sum(r.row_count),0) AS rows
               FROM me_datasets d LEFT JOIN me_resources r ON r.dataset_id=d.id
               GROUP BY d.id, d.title ORDER BY rows DESC LIMIT 15"""
        )
        top = cur.fetchall()
        cur.execute(
            "SELECT count(*) FROM me_resources WHERE import_status='error'"
        )
        n_err = cur.fetchone()[0]

    print("\n" + "=" * 64)
    print("  me.gov.ua open-data mirror — verification")
    print("=" * 64)
    print(f"  datasets                : {n_datasets}")
    print(f"  resources (stored/CKAN) : {n_res} / {expected_res}")
    print(f"  records (me_records)    : {n_records:,}")
    print(f"  resources in error      : {n_err}")
    print("\n  resources by import_status:")
    for st, c in by_status:
        print(f"    {st:12s} {c}")
    print("\n  resources by format:")
    for f, c in by_format:
        print(f"    {f:12s} {c}")
    print("\n  top datasets by imported rows:")
    for title, res, rows in top:
        t = (title or "")[:50]
        print(f"    {rows:>10,}  ({res} res)  {t}")
    print("=" * 64)

    problems = []
    if n_datasets == 0:
        problems.append("no datasets — run discover")
    if n_res < expected_res:
        problems.append(f"resource gap: stored {n_res} < CKAN {expected_res}")
    if n_err:
        problems.append(f"{n_err} resources in error state")
    if problems:
        print("  ⚠ issues:")
        for p in problems:
            print(f"    - {p}")
        return 1
    print("  ✓ all checks passed")
    return 0


# ── main ────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="me.gov.ua open-data mirror pipeline")
    ap.add_argument(
        "stage",
        choices=["discover", "download", "import", "verify", "all"],
    )
    ap.add_argument("--limit", type=int, default=None,
                    help="download: max resources to fetch this run")
    ap.add_argument("--reimport", action="store_true",
                    help="import: re-parse already-imported resources")
    args = ap.parse_args()

    log.info("target DB: %s:%s/%s (staging=%s)",
             os.environ.get("POSTGRES_HOST", "localhost"),
             os.environ.get("POSTGRES_PORT", "5435"),
             os.environ.get("POSTGRES_DB", "openreyestr"),
             STAGING_DIR)

    conn = connect()
    try:
        if args.stage in ("discover", "all"):
            discover(conn)
        if args.stage in ("download", "all"):
            download(conn, args.limit)
        if args.stage in ("import", "all"):
            import_records(conn, args.reimport)
        if args.stage in ("verify", "all"):
            rc = verify(conn)
            sys.exit(rc)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
