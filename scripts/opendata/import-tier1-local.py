#!/usr/bin/env python3
"""
Import TIER 1 data.gov.ua datasets into local PostgreSQL.
Reads files from /mnt/opendata/ (or --data-dir), inserts into DB.

Usage:
    python3 import-tier1-local.py                          # import all
    python3 import-tier1-local.py --dataset passports      # single dataset
    python3 import-tier1-local.py --dataset vehicles --data-dir /mnt/opendata
    python3 import-tier1-local.py --list

Requires: pip install psycopg2-binary openpyxl
"""

import argparse
import csv
import io
import json
import logging
import os
import sys
import time
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import psycopg2
from psycopg2.extras import execute_values

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

DATA_DIR = Path(os.environ.get("DATA_DIR", "/mnt/opendata"))
BATCH_SIZE = 5000

def get_db():
    return psycopg2.connect(
        host=os.environ.get("POSTGRES_HOST", "127.0.0.1"),
        port=int(os.environ.get("POSTGRES_PORT", "5432")),
        user=os.environ.get("POSTGRES_USER", "secondlayer"),
        password=os.environ.get("POSTGRES_PASSWORD", "secondlayer"),
        dbname=os.environ.get("POSTGRES_DB", "secondlayer_local"),
    )


# ── Passports (invalid) ────────────────────────────────────
def import_passports(data_dir: Path, table: str = "opendata_invalid_passports"):
    folder = data_dir / ("passports-invalid" if "foreign" not in table else "passports-foreign-invalid")
    if not folder.exists():
        log.warning(f"Folder not found: {folder}")
        return

    conn = get_db()
    cur = conn.cursor()
    total = 0

    json_files = sorted(f for f in folder.iterdir() if f.suffix == ".json" and "shema" not in f.name.lower())
    log.info(f"[{table}] Found {len(json_files)} JSON files in {folder}")

    for jf in json_files:
        log.info(f"  Processing {jf.name} ({jf.stat().st_size / 1024 / 1024:.1f} MB)...")
        with open(jf, "r", encoding="utf-8-sig") as f:
            records = json.load(f)

        batch = []
        for r in records:
            theft_date = (r.get("THEFT_DATA") or "")[:10] or None
            insert_date = (r.get("INSERT_DATE") or "")[:10] or None
            batch.append((
                r.get("ID"), r.get("OVD"), r.get("D_SERIES"), r.get("D_NUMBER"),
                r.get("D_TYPE"), r.get("D_STATUS"), theft_date, insert_date,
            ))
            if len(batch) >= BATCH_SIZE:
                execute_values(cur, f"""
                    INSERT INTO {table} (id, ovd, d_series, d_number, d_type, d_status, theft_date, insert_date)
                    VALUES %s ON CONFLICT (id) DO NOTHING
                """, batch)
                conn.commit()
                total += len(batch)
                batch = []

        if batch:
            execute_values(cur, f"""
                INSERT INTO {table} (id, ovd, d_series, d_number, d_type, d_status, theft_date, insert_date)
                VALUES %s ON CONFLICT (id) DO NOTHING
            """, batch)
            conn.commit()
            total += len(batch)

        log.info(f"    {len(records)} records, running total: {total}")

    cur.close()
    conn.close()
    log.info(f"[{table}] Done: {total} records imported")


# ── Vehicles ────────────────────────────────────────────────
def import_vehicles(data_dir: Path):
    folder = data_dir / "vehicles"
    if not folder.exists():
        log.warning(f"Folder not found: {folder}")
        return

    conn = get_db()
    cur = conn.cursor()
    total = 0

    zip_files = sorted(f for f in folder.iterdir() if f.suffix == ".zip")
    log.info(f"[vehicles] Found {len(zip_files)} ZIP files")

    for zf in zip_files:
        log.info(f"  Processing {zf.name} ({zf.stat().st_size / 1024 / 1024:.1f} MB)...")
        with zipfile.ZipFile(zf) as z:
            for csv_name in z.namelist():
                if not csv_name.lower().endswith(".csv"):
                    continue
                with z.open(csv_name) as f:
                    text = io.TextIOWrapper(f, encoding="utf-8", errors="replace")
                    reader = csv.DictReader(text, delimiter=";", quotechar='"')
                    batch = []
                    file_count = 0

                    for row in reader:
                        d_reg = row.get("D_REG", "")
                        # Parse DD.MM.YY or DD.MM.YYYY
                        d_reg_parsed = None
                        if d_reg:
                            try:
                                parts = d_reg.split(".")
                                if len(parts) == 3:
                                    day, month, year = parts
                                    if len(year) == 2:
                                        year = "20" + year if int(year) < 50 else "19" + year
                                    d_reg_parsed = f"{year}-{month}-{day}"
                            except:
                                pass

                        make_year = None
                        try:
                            make_year = int(row.get("MAKE_YEAR", ""))
                        except:
                            pass

                        capacity = None
                        try:
                            capacity = int(row.get("CAPACITY", ""))
                        except:
                            pass

                        own_weight = None
                        try:
                            own_weight = int(row.get("OWN_WEIGHT", ""))
                        except:
                            pass

                        total_weight = None
                        try:
                            total_weight = int(row.get("TOTAL_WEIGHT", ""))
                        except:
                            pass

                        batch.append((
                            row.get("PERSON"), row.get("REG_ADDR_KOATUU"), row.get("OPER_CODE"),
                            row.get("OPER_NAME"), d_reg_parsed, row.get("DEP_CODE"), row.get("DEP"),
                            row.get("BRAND"), row.get("MODEL"), row.get("VIN"),
                            make_year, row.get("COLOR"), row.get("KIND"), row.get("BODY"),
                            row.get("PURPOSE"), row.get("FUEL"), capacity, own_weight, total_weight,
                            row.get("N_REG_NEW"),
                        ))

                        if len(batch) >= BATCH_SIZE:
                            execute_values(cur, """
                                INSERT INTO opendata_vehicle_registrations
                                    (person_type, reg_addr_koatuu, oper_code, oper_name, d_reg,
                                     dep_code, dep, brand, model, vin, make_year, color, kind,
                                     body, purpose, fuel, capacity, own_weight, total_weight, n_reg_new)
                                VALUES %s
                            """, batch)
                            conn.commit()
                            total += len(batch)
                            file_count += len(batch)
                            batch = []

                    if batch:
                        execute_values(cur, """
                            INSERT INTO opendata_vehicle_registrations
                                (person_type, reg_addr_koatuu, oper_code, oper_name, d_reg,
                                 dep_code, dep, brand, model, vin, make_year, color, kind,
                                 body, purpose, fuel, capacity, own_weight, total_weight, n_reg_new)
                            VALUES %s
                        """, batch)
                        conn.commit()
                        total += len(batch)
                        file_count += len(batch)

                    log.info(f"    {csv_name}: {file_count} rows, running total: {total}")

    cur.close()
    conn.close()
    log.info(f"[vehicles] Done: {total} records imported")


# ── Terrorism persons ───────────────────────────────────────
def import_terrorism(data_dir: Path):
    folder = data_dir / "dsfmu-terrorism"
    if not folder.exists():
        log.warning(f"Folder not found: {folder}")
        return

    conn = get_db()
    cur = conn.cursor()
    total = 0

    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl required: pip install openpyxl")
        return

    xlsx_files = sorted(f for f in folder.iterdir() if f.suffix in (".xlsx", ".xls"))
    # Use the latest file
    for xf in xlsx_files:
        log.info(f"  Processing {xf.name}...")
        try:
            wb = openpyxl.load_workbook(xf, read_only=True)
            ws = wb.active
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c or "").strip() for c in row]
                    continue
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if row[i]}
                if not row_dict:
                    continue
                # Try to extract key fields
                name_ua = str(row[1] if len(row) > 1 and row[1] else "")
                name_en = str(row[2] if len(row) > 2 and row[2] else "")
                cur.execute("""
                    INSERT INTO opendata_terrorism_persons (row_data, name_ua, name_en)
                    VALUES (%s, %s, %s)
                """, (json.dumps(row_dict, ensure_ascii=False, default=str), name_ua, name_en))
                total += 1
            wb.close()
            conn.commit()
            log.info(f"    {total} records")
            break  # Use only latest
        except Exception as e:
            log.error(f"    Error: {e}")
            continue

    cur.close()
    conn.close()
    log.info(f"[terrorism] Done: {total} records imported")


# ── Terrorism organizations ─────────────────────────────────
def import_terrorism_orgs(data_dir: Path):
    folder = data_dir / "dsfmu-terrorism-orgs"
    if not folder.exists():
        log.warning(f"Folder not found: {folder}")
        return

    conn = get_db()
    cur = conn.cursor()
    total = 0

    csv_files = [f for f in folder.iterdir() if f.suffix == ".csv"]
    for cf in csv_files:
        log.info(f"  Processing {cf.name}...")
        with open(cf, "r", encoding="windows-1251", errors="replace") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row in reader:
                cur.execute("""
                    INSERT INTO opendata_terrorism_orgs
                        (name_ua, name_ua_short, name_en, head_name_ua, head_name_en,
                         head_dob, org_country, report_date, report_num, raw_data)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    row.get("nameUA"), row.get("nameUAShort"), row.get("nameEn"),
                    row.get("headNameUA"), row.get("headNameEn"),
                    row.get("headDateOfBirth"), row.get("orgCountry"),
                    row.get("raportDate"), row.get("raportNum"),
                    json.dumps(row, ensure_ascii=False),
                ))
                total += 1
        conn.commit()

    cur.close()
    conn.close()
    log.info(f"[terrorism_orgs] Done: {total} records imported")


# ── Lustration ──────────────────────────────────────────────
def import_lustration(data_dir: Path):
    folder = data_dir / "lustration"
    if not folder.exists():
        log.warning(f"Folder not found: {folder}")
        return

    conn = get_db()
    cur = conn.cursor()
    total = 0

    zip_files = [f for f in folder.iterdir() if f.suffix == ".zip"]
    for zf in zip_files:
        log.info(f"  Processing {zf.name}...")
        with zipfile.ZipFile(zf) as z:
            for xml_name in z.namelist():
                if not xml_name.lower().endswith(".xml"):
                    continue
                with z.open(xml_name) as f:
                    raw = f.read()
                    text = raw.decode("windows-1251", errors="replace")
                    root = ET.fromstring(text)
                    for record in root.findall("RECORD"):
                        fio = (record.findtext("FIO") or "").strip()
                        job = (record.findtext("JOB") or "").strip()
                        judgment = (record.findtext("JUDGMENT_COMPOSITION") or "").strip()
                        period = (record.findtext("PERIOD") or "").strip()
                        if fio:
                            cur.execute("""
                                INSERT INTO opendata_lustration (fio, job, judgment, period)
                                VALUES (%s, %s, %s, %s)
                            """, (fio, job, judgment, period))
                            total += 1
        conn.commit()

    cur.close()
    conn.close()
    log.info(f"[lustration] Done: {total} records imported")


# ── State aid (AMKU) ────────────────────────────────────────
def import_state_aid(data_dir: Path):
    folder = data_dir / "state-aid-amku"
    if not folder.exists():
        log.warning(f"Folder not found: {folder}")
        return

    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl required: pip install openpyxl")
        return

    conn = get_db()
    cur = conn.cursor()
    total = 0

    # Use the latest file (by name sort)
    xlsx_files = sorted(f for f in folder.iterdir() if f.suffix == ".xlsx")
    if not xlsx_files:
        log.warning("No XLSX files found")
        return

    latest = xlsx_files[-1]
    log.info(f"  Processing {latest.name}...")
    wb = openpyxl.load_workbook(latest, read_only=True)
    ws = wb.active
    headers = None

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c or "").strip() for c in row]
            continue
        if not any(row):
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and val is not None:
                row_dict[headers[i]] = val

        cur.execute("""
            INSERT INTO opendata_state_aid (row_data, provider_name, recipient_name, program_name)
            VALUES (%s, %s, %s, %s)
        """, (
            json.dumps(row_dict, ensure_ascii=False, default=str),
            str(row_dict.get(headers[1], "")) if len(headers) > 1 else None,  # provider
            str(row_dict.get(headers[3], "")) if len(headers) > 3 else None,  # recipient
            str(row_dict.get(headers[5], "")) if len(headers) > 5 else None,  # program
        ))
        total += 1

    conn.commit()
    wb.close()
    cur.close()
    conn.close()
    log.info(f"[state_aid] Done: {total} records imported")


# ── Financial statements (summary only — header extraction) ─
def import_finzvit(data_dir: Path):
    folder = data_dir / "finzvit"
    if not folder.exists():
        log.warning(f"Folder not found: {folder}")
        return

    conn = get_db()
    cur = conn.cursor()
    total = 0

    # Process latest year file first (fin_zvit_2025)
    zip_files = sorted(
        (f for f in folder.iterdir() if f.suffix == ".zip" and "opis_" not in f.name.lower()),
        key=lambda f: f.stat().st_size,
        reverse=True,
    )

    for outer_zip in zip_files:
        log.info(f"  Processing {outer_zip.name} ({outer_zip.stat().st_size / 1024 / 1024:.0f} MB)...")
        file_count = 0
        try:
            with zipfile.ZipFile(outer_zip) as oz:
                for inner_name in oz.namelist():
                    if inner_name.lower().endswith(".zip"):
                        # Nested ZIP containing XMLs
                        with oz.open(inner_name) as inner_file:
                            inner_data = inner_file.read()
                            try:
                                with zipfile.ZipFile(io.BytesIO(inner_data)) as iz:
                                    for xml_name in iz.namelist():
                                        if not xml_name.lower().endswith(".xml"):
                                            continue
                                        try:
                                            with iz.open(xml_name) as xf:
                                                xml_content = xf.read().decode("utf-8", errors="replace")
                                                root = ET.fromstring(xml_content)
                                                head = root.find("DECLARHEAD")
                                                if head is None:
                                                    continue
                                                tin = head.findtext("TIN", "")
                                                if not tin:
                                                    continue
                                                # Extract form type from filename
                                                form_type = ""
                                                parts = xml_name.split("S")
                                                if len(parts) >= 2:
                                                    form_type = "S" + parts[-1].split(".")[0].split("_")[0][:7]

                                                cur.execute("""
                                                    INSERT INTO opendata_financial_statements
                                                        (tin, c_doc, c_doc_sub, c_doc_ver, period_year, period_month,
                                                         period_type, c_reg, c_raj, form_type, raw_xml)
                                                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                                                    ON CONFLICT (tin, period_year, form_type, c_doc_sub) DO NOTHING
                                                """, (
                                                    tin,
                                                    head.findtext("C_DOC"),
                                                    head.findtext("C_DOC_SUB"),
                                                    head.findtext("C_DOC_VER"),
                                                    int(head.findtext("PERIOD_YEAR") or 0) or None,
                                                    int(head.findtext("PERIOD_MONTH") or 0) or None,
                                                    int(head.findtext("PERIOD_TYPE") or 0) or None,
                                                    head.findtext("C_REG"),
                                                    head.findtext("C_RAJ"),
                                                    form_type,
                                                    xml_content,
                                                ))
                                                file_count += 1
                                                total += 1

                                                if file_count % BATCH_SIZE == 0:
                                                    conn.commit()
                                                    log.info(f"    {file_count} XMLs processed, total: {total}")
                                        except ET.ParseError:
                                            pass
                            except zipfile.BadZipFile:
                                pass
                    elif inner_name.lower().endswith(".xml"):
                        # Direct XML in outer ZIP
                        try:
                            with oz.open(inner_name) as xf:
                                xml_content = xf.read().decode("utf-8", errors="replace")
                                root = ET.fromstring(xml_content)
                                head = root.find("DECLARHEAD")
                                if head is None:
                                    continue
                                tin = head.findtext("TIN", "")
                                if not tin:
                                    continue
                                form_type = ""
                                parts = inner_name.split("S")
                                if len(parts) >= 2:
                                    form_type = "S" + parts[-1].split(".")[0].split("_")[0][:7]
                                cur.execute("""
                                    INSERT INTO opendata_financial_statements
                                        (tin, c_doc, c_doc_sub, c_doc_ver, period_year, period_month,
                                         period_type, c_reg, c_raj, form_type, raw_xml)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                                    ON CONFLICT (tin, period_year, form_type, c_doc_sub) DO NOTHING
                                """, (
                                    tin,
                                    head.findtext("C_DOC"),
                                    head.findtext("C_DOC_SUB"),
                                    head.findtext("C_DOC_VER"),
                                    int(head.findtext("PERIOD_YEAR") or 0) or None,
                                    int(head.findtext("PERIOD_MONTH") or 0) or None,
                                    int(head.findtext("PERIOD_TYPE") or 0) or None,
                                    head.findtext("C_REG"),
                                    head.findtext("C_RAJ"),
                                    form_type,
                                    xml_content,
                                ))
                                file_count += 1
                                total += 1
                        except ET.ParseError:
                            pass

            conn.commit()
            log.info(f"    {outer_zip.name}: {file_count} XMLs, running total: {total}")
        except Exception as e:
            log.error(f"    Error processing {outer_zip.name}: {e}")
            conn.rollback()

    cur.close()
    conn.close()
    log.info(f"[finzvit] Done: {total} records imported")


# ── Main ────────────────────────────────────────────────────
IMPORTERS = {
    "passports": lambda d: import_passports(d, "opendata_invalid_passports"),
    "passports_foreign": lambda d: import_passports(d, "opendata_invalid_passports_foreign"),
    "vehicles": import_vehicles,
    "terrorism": import_terrorism,
    "terrorism_orgs": import_terrorism_orgs,
    "lustration": import_lustration,
    "state_aid": import_state_aid,
    "finzvit": import_finzvit,
}

def main():
    parser = argparse.ArgumentParser(description="Import TIER 1 datasets into local PostgreSQL")
    parser.add_argument("--dataset", type=str, help="Dataset key (comma-separated)")
    parser.add_argument("--data-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--list", action="store_true")
    args = parser.parse_args()

    if args.list:
        for k in IMPORTERS:
            log.info(f"  {k}")
        return

    data_dir = args.data_dir
    if not data_dir.exists():
        log.error(f"Data dir not found: {data_dir}")
        sys.exit(1)

    # Run migration first
    log.info("Running migration 117...")
    migration = Path(__file__).parent.parent.parent / "mcp_backend" / "src" / "migrations" / "117_tier1_opendata_tables.sql"
    if migration.exists():
        conn = get_db()
        cur = conn.cursor()
        cur.execute(migration.read_text())
        conn.commit()
        cur.close()
        conn.close()
        log.info("  Migration applied")

    datasets = list(IMPORTERS.keys())
    if args.dataset:
        datasets = [k.strip() for k in args.dataset.split(",")]

    log.info(f"Importing {len(datasets)} datasets from {data_dir}")
    start = time.time()

    for ds in datasets:
        if ds not in IMPORTERS:
            log.error(f"Unknown dataset: {ds}")
            continue
        log.info(f"\n{'=' * 60}")
        log.info(f"  [{ds}]")
        log.info(f"{'=' * 60}")
        ds_start = time.time()
        IMPORTERS[ds](data_dir)
        log.info(f"  Elapsed: {time.time() - ds_start:.1f}s")

    log.info(f"\n=== All done in {time.time() - start:.1f}s ===")


if __name__ == "__main__":
    main()
