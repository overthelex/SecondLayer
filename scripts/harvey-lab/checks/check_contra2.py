#!/usr/bin/env python3
"""Verify every contradiction criterion is actually grounded in the documents.

The previous checker matched English criterion prose against Ukrainian
documents and drowned in false positives, which is worse than no check: real
gaps become indistinguishable from phantoms.

This one only asserts what is language-independent:
  * every amount (4+ digit number) cited in the criterion appears in some
    document, in either "1 234 567" or "1234567" form
  * every dd.mm.yyyy date cited appears in some document
  * every case / proceeding number cited appears in some document

Prose is ignored on purpose. If a criterion's values are all present, the
conflict it describes is available to the reader; whether the wording is fair is
a question for the lawyer, not for a script.
"""

import json
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

NUM = re.compile(r"\b\d[\d  ]{3,}\d\b")
DATE = re.compile(r"\b\d{2}\.\d{2}\.\d{4}\b")
CASE = re.compile(r"\b\d{3}/\d{3,4}/\d{2}\b")


def doc_text(p: Path) -> str:
    if p.suffix == ".xlsx":
        ws = load_workbook(p).active
        return " ".join(str(c) for r in ws.iter_rows(values_only=True)
                        for c in r if c is not None)
    if p.suffix == ".docx":
        xml = zipfile.ZipFile(p).read("word/document.xml").decode()
        return " ".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml, re.S))
    return p.read_text(encoding="utf-8", errors="ignore")


def norm(s: str) -> str:
    return s.replace(" ", "").replace(" ", "")


def main():
    root = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("tasks")
    bad = total = 0
    for tj in sorted(root.glob("diligence/*-v2/task.json")):
        t = tj.parent
        corpus = " ".join(doc_text(p) for p in sorted((t / "documents").iterdir()))
        flat = norm(corpus)
        cfg = json.loads(tj.read_text(encoding="utf-8"))
        rows = []
        for c in cfg["criteria"]:
            if not c["title"].startswith("Resolves contradiction"):
                continue
            total += 1
            m = c["match_criteria"]
            want = set(NUM.findall(m)) | set(DATE.findall(m)) | set(CASE.findall(m))
            missing = [v for v in want
                       if v not in corpus and norm(v) not in flat]
            if missing:
                bad += 1
                rows.append((c["id"], c["title"], missing))
        if rows:
            print(f"{t.name}")
            for cid, title, miss in rows:
                print(f"  UNGROUNDED {cid} {title[:52]}")
                print(f"             values absent from documents: {miss}")
    print()
    print(f"contradiction criteria checked: {total}, ungrounded: {bad}")
    sys.exit(1 if bad else 0)


if __name__ == "__main__":
    main()
