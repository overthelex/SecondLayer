"""Each value a contradiction criterion asks the memo to DERIVE must be arithmetically right."""
import json, re, zipfile, sys
from pathlib import Path
from openpyxl import load_workbook

def nums_in(p):
    if p.suffix == ".xlsx":
        ws = load_workbook(p).active
        return {int(c) for r in ws.iter_rows(values_only=True) for c in r
                if isinstance(c, (int, float)) and float(c).is_integer() and c >= 1000}
    if p.suffix == ".docx":
        xml = zipfile.ZipFile(p).read("word/document.xml").decode()
        t = " ".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml, re.S))
    else:
        t = p.read_text(encoding="utf-8", errors="ignore")
    return {int(re.sub(r"\D", "", m)) for m in re.findall(r"\b\d[\d  ]{3,}\d\b", t)}

root = Path(sys.argv[1])
bad = 0
for tj in sorted(root.glob("diligence/*-v2/task.json")):
    t = tj.parent
    present = set()
    for p in (t / "documents").iterdir():
        present |= nums_in(p)
    cfg = json.loads(tj.read_text(encoding="utf-8"))
    for c in cfg["criteria"]:
        if not c["title"].startswith("Resolves contradiction"):
            continue
        cited = {int(re.sub(r"\D", "", v)) for v in re.findall(r"\b\d[\d  ]{3,}\d\b", c["match_criteria"])}
        derived = cited - present
        for d in sorted(derived):
            # is it a difference of two present values, or their sum?
            diff = any(abs(a - b) == d for a in present for b in present if a != b)
            summ = (d == sum(x for x in present if x in present and False)) or False
            ok = diff or d in {sum(s) for s in [tuple(present)]}
            explain = "difference of two document values" if diff else ("sum" if ok else "UNEXPLAINED")
            if not (diff or ok):
                bad += 1
            print(f"{t.name:32s} {c['id']} derived {d:>10,} -> {explain}")
print(f"\nunexplained derived values: {bad}")
sys.exit(1 if bad else 0)
