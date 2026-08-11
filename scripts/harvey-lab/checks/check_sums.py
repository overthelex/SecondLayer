"""Verify every computed exposure total in a criterion against its source xlsx."""
import json, re, glob, sys
from pathlib import Path
from openpyxl import load_workbook

bad = 0
checked = 0
for tj in sorted(glob.glob("tasks/**/task.json", recursive=True)):
    cfg = json.load(open(tj, encoding="utf-8"))
    if cfg.get("jurisdiction") != "UA":
        continue
    vp = Path(tj).parent / "documents" / "vykonavchi-provadzhennya.xlsx"
    if not vp.exists():
        continue
    ws = load_workbook(vp).active
    hdr = [c.value for c in ws[1]]
    i_amt, i_state = hdr.index("Сума, грн"), hdr.index("Стан")
    open_sum = sum(r[i_amt] for r in ws.iter_rows(min_row=2, values_only=True)
                   if r[i_state] == "відкрито")
    crit = next((c for c in cfg["criteria"]
                 if "open enforcement" in c["title"].lower()), None)
    if not crit:
        continue
    m = re.search(r"as ([\d  ]+) UAH", crit["match_criteria"])
    stated = int(re.sub(r"\D", "", m.group(1))) if m else None
    checked += 1
    ok = stated == open_sum
    bad += (not ok)
    print(f"{'OK ' if ok else 'BAD'} {Path(tj).parent.name:42s} "
          f"stated={stated:>10,} computed={open_sum:>10,}")
print(f"\nchecked {checked}, mismatches {bad}")
sys.exit(1 if bad else 0)
