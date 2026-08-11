"""Every value a contradiction criterion cites must actually appear in a document."""
import json, re, zipfile, glob, sys
from pathlib import Path
from openpyxl import load_workbook

def text_of(p):
    if p.suffix == ".docx":
        xml = zipfile.ZipFile(p).read("word/document.xml").decode()
        return " ".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml, re.S))
    if p.suffix == ".xlsx":
        ws = load_workbook(p).active
        return " ".join(str(c) for r in ws.iter_rows(values_only=True) for c in r if c is not None)
    return p.read_text(encoding="utf-8", errors="ignore")

bad = 0
for tj in sorted(glob.glob("tasks/diligence/*-v2/task.json")):
    t = Path(tj).parent
    corpus = " ".join(text_of(p) for p in sorted((t / "documents").iterdir()))
    norm = corpus.replace(" ", " ")
    print("=" * 78); print(t.name)
    cfg = json.loads(Path(tj).read_text(encoding="utf-8"))
    for c in cfg["criteria"]:
        if not c["title"].startswith("Resolves contradiction"):
            continue
        # pull the quoted/parenthesised values out of the criterion text
        vals = re.findall(r"gives ([^;)]+?)[;)]", c["match_criteria"])
        missing = []
        for v in vals:
            v = v.strip().strip("'\"")
            probe = v.replace(" UAH", "").replace("'", "").strip()
            if probe and probe not in norm and probe.replace(" ", "") not in norm.replace(" ", ""):
                missing.append(v)
        status = "OK " if not missing else "MISSING"
        if missing: bad += 1
        print(f"  {status} {c['title'][:60]}")
        for m in missing:
            print(f"        not in documents: {m!r}")
print(f"\ncontradiction criteria with unsupported values: {bad}")
sys.exit(1 if bad else 0)
