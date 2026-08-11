#!/usr/bin/env python3
"""Audit the reviewer package as a shipped artifact, not as an intention.

Checks, in order of how badly each would embarrass us:
 1. every ЄДРПОУ-shaped code in the documents FAILS its checksum, because the
    letter states outright that none can collide with a real company
 2. the xlsx carries the corrected criterion text, not the pre-fix wording
 3. every criterion in the xlsx matches the repo verbatim
 4. each stage folder holds exactly the documents of its task, plus ЗАВДАННЯ.txt
 5. counts in the plan sheet agree with the per-stage sheets
"""

import json
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

PKG = Path(sys.argv[1])
TASKS = Path(sys.argv[2])
XLSX = PKG / "ВАЛІДАЦІЯ-критеріїв.xlsx"
fail = 0


def problem(msg):
    global fail
    fail += 1
    print("  PROBLEM:", msg)


def edrpou_ok(d: str) -> bool:
    if len(set(d)) <= 2:
        return False
    n = [int(c) for c in d]
    base = [1, 2, 3, 4, 5, 6, 7] if int(d) < 30000000 else [7, 1, 2, 3, 4, 5, 6]
    c = sum(w * x for w, x in zip(base, n[:7])) % 11
    if c >= 10:
        c = sum(w * x for w, x in zip([w + 2 for w in base], n[:7])) % 11
        if c >= 10:
            return False
    return c == n[7]


def doc_text(p: Path) -> str:
    if p.suffix in (".docx", ".xlsx"):
        try:
            z = zipfile.ZipFile(p)
        except zipfile.BadZipFile:
            return ""
        out = []
        for n in z.namelist():
            if n.endswith(".xml"):
                x = z.read(n).decode("utf-8", "ignore")
                out += re.findall(r"<(?:w|a):t[^>]*>(.*?)</(?:w|a):t>", x, re.S)
                out += re.findall(r"<t[^>]*>(.*?)</t>", x, re.S)
                out += re.findall(r"<v>(.*?)</v>", x, re.S)
        return " ".join(out)
    return p.read_text(encoding="utf-8", errors="ignore")


print("1. ЄДРПОУ codes in package documents must all FAIL the checksum")
real = []
for f in sorted(PKG.rglob("*")):
    if not f.is_file() or f.name.endswith(".xlsx") and f.parent == PKG:
        continue
    for m in re.finditer(r"\b\d{8}\b", doc_text(f)):
        if edrpou_ok(m.group()):
            real.append((f.relative_to(PKG), m.group()))
if real:
    for p, c in real[:10]:
        problem(f"code {c} passes the checksum, in {p}")
else:
    print("   ok: no code in any document validates")

print("2/3. xlsx criterion text matches the repo, including the fixes")
# index the repo once; rglob-per-row was O(rows x 2030 files)
BY_TITLE = {}
for p_ in TASKS.rglob("task.json"):
    try:
        c = json.loads(p_.read_text(encoding="utf-8"))
    except Exception:
        continue
    BY_TITLE[c.get("title")] = (p_, {x["id"]: x for x in c.get("criteria", [])})
DIRS = {p_.parent.name: p_.parent for p_ in TASKS.rglob("task.json")}
wb = load_workbook(XLSX)
seen = 0
for sheet in [s for s in wb.sheetnames if s.startswith("Етап")]:
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        _, title, cid, _, match, *_ = row
        seen += 1
        entry = BY_TITLE.get(title)
        if entry is None:
            problem(f"{sheet} {cid}: task titled {title!r} not found in repo")
            continue
        src_path, crit_by_id = entry
        cr = crit_by_id.get(cid)
        if cr is None:
            problem(f"{sheet} {cid}: criterion absent from {src_path}")
        elif cr["match_criteria"] != match:
            problem(f"{sheet} {cid}: text differs from repo")
print(f"   checked {seen} criteria rows")

print("4. stage folders hold their task's documents")
for sheet in [s for s in wb.sheetnames if s.startswith("Етап")]:
    no = sheet.split()[-1]
    d = PKG / f"етап-{no}"
    if not d.is_dir():
        problem(f"missing folder етап-{no}")
        continue
    for task_dir in d.iterdir():
        if not task_dir.is_dir():
            continue
        src = DIRS.get(task_dir.name)
        if src is None:
            problem(f"етап-{no}/{task_dir.name}: no matching task in repo")
            continue
        want = {p.name for p in (src / "documents").iterdir()}
        got = {p.name for p in task_dir.iterdir()} - {"ЗАВДАННЯ.txt"}
        if want != got:
            problem(f"етап-{no}/{task_dir.name}: docs differ "
                    f"missing={sorted(want - got)} extra={sorted(got - want)}")
        if not (task_dir / "ЗАВДАННЯ.txt").exists():
            problem(f"етап-{no}/{task_dir.name}: ЗАВДАННЯ.txt missing")

print("5. plan sheet totals agree with the stage sheets")
plan = wb["План"]
rows = [r for r in plan.iter_rows(min_row=2, values_only=True) if r and r[0] not in (None, "РАЗОМ")]
for no, _, _, n_crit, *_ in rows:
    ws = wb[f"Етап {no}"]
    actual = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[2])
    if actual != n_crit:
        problem(f"етап {no}: plan says {n_crit} criteria, sheet has {actual}")
total_plan = sum(r[3] for r in rows)
print(f"   plan total {total_plan} criteria across {len(rows)} stages")

print()
print("PACKAGE CLEAN" if not fail else f"{fail} PROBLEM(S) FOUND")
sys.exit(1 if fail else 0)
