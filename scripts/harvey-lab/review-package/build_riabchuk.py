#!/usr/bin/env python3
"""Build the manual-validation package for адвокат О. Рябчук.

Scope is deliberately narrow: only the tasks that are stable. The 14 v1
register-screening tasks are being rebuilt and are excluded, so no reviewer time
is spent on criteria that will be deleted.

Stages are sized so each is about an hour. The fixed cost per task is reading
its documents once, so a task is normally one stage; the 25-criterion flagship
is split in two.

Usage:
    uv run --with openpyxl python build_riabchuk.py <tasks_root> <out_dir>
"""

import json
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

LIT = "litigation-dispute-resolution"
DIL = "diligence"

# (stage no, label, [(task_path, criterion slice or None)], minutes)
STAGES = [
    (1, "Позовна давність під час воєнного стану, частина 1",
     [(f"{LIT}/ua-limitation-period-martial-law", (0, 13))], 60),
    (2, "Позовна давність під час воєнного стану, частина 2",
     [(f"{LIT}/ua-limitation-period-martial-law", (13, None))], 50),
    (3, "Стаття 259 ЦК: договірна зміна строку",
     [(f"{LIT}/ua-limitation-contractual-shortening-void", None),
      (f"{LIT}/ua-limitation-extended-by-agreement", None)], 60),
    (4, "Заява про давність і спеціальний однорічний строк",
     [(f"{LIT}/ua-limitation-not-raised-by-party", None),
      (f"{LIT}/ua-limitation-penalty-one-year", None)], 60),
    (5, "Карантин проти воєнного стану: продовження і зупинення",
     [(f"{LIT}/ua-limitation-quarantine-vs-martial-law", None)], 45),
    (6, "Реєстрова перевірка: агротрейдинг",
     [(f"{DIL}/ua-agro-supply-v2", None)], 60),
    (7, "Реєстрова перевірка: перевезення",
     [(f"{DIL}/ua-transport-fleet-v2", None)], 60),
    (8, "Реєстрова перевірка: харчова галузь",
     [(f"{DIL}/ua-food-tax-v2", None)], 60),
]

HDR = ["Етап", "Задача", "ID", "Що перевіряє критерій (EN)",
       "Повне формулювання для судді (EN)", "Тип", "ВЕРДИКТ",
       "Коментар / правильне формулювання"]

GUIDE = [
    ("Що це і навіщо", True),
    ("", False),
    ("Ми додаємо українські задачі до міжнародного бенчмарку для оцінки ШІ на "
     "юридичній роботі (Harvey Legal Agent Benchmark). Кожна задача це справжня "
     "робоча ситуація: комплект документів і завдання підготувати меморандум. "
     "Відповідь ШІ оцінюється за списком критеріїв.", False),
    ("", False),
    ("Ваше завдання: перевірити, чи кожен критерій юридично коректний. Не треба "
     "оцінювати роботу ШІ. Треба оцінити САМ КРИТЕРІЙ.", False),
    ("", False),
    ("Як влаштований критерій", True),
    ("", False),
    ("Критерій бінарний: або виконано, або ні, півбалів немає. Задача "
     "зараховується лише якщо пройдені ВСІ критерії. Тому жоден критерій не "
     "може бути спірним, оціночним або таким, де можливі дві правильні думки.", False),
    ("", False),
    ("Формулювання завжди у вигляді PASS if ... FAIL if ... англійською. "
     "Англійська тут навмисно: критерії читатимуть іноземні мейнтейнери, які не "
     "знають української. Самі документи і меморандум українською.", False),
    ("", False),
    ("Колонка ВЕРДИКТ", True),
    ("", False),
    ("OK       — юридично правильно і однозначно, залишаємо як є", False),
    ("ПРАВКА   — по суті вірно, але формулювання треба уточнити. Напишіть у "
     "коментарі, як саме.", False),
    ("ПОМИЛКА  — юридично неправильно. Критерій треба прибрати або переписати. "
     "Поясніть чому.", False),
    ("", False),
    ("На що звертати особливу увагу", True),
    ("", False),
    ("1. Тип «oracle» означає, що критерій має перевірятися механічно за "
     "текстом закону або за документом. Там будь-яка неоднозначність це "
     "помилка.", False),
    ("2. Критерій не повинен вимагати конкретного формулювання. Він має "
     "перевіряти зміст. Якщо критерій пройде лише за умови певної фрази, це "
     "ПРАВКА.", False),
    ("3. Критерій не повинен виконуватись порожнім меморандумом. Якщо критерій "
     "сформульований через заперечення («не стверджує», «не вигадує») і "
     "порожній документ його задовольняє, це ПРАВКА.", False),
    ("", False),
    ("Важливо про дати", True),
    ("", False),
    ("Справи у літигаційних задачах умисно датовані ДО 14.05.2025. "
     "Застосовувати треба редакцію закону, чинну на дату рішення суду першої "
     "інстанції, яка вказана в задачі.", False),
    ("Пункт 19 Прикінцевих та перехідних положень ЦК (зупинення перебігу "
     "позовної давності на час воєнного стану) був чинним на ті дати і "
     "виключений лише 14.05.2025 Законом № 4434-IX. Посилання на нього це не "
     "помилка, а суть задачі.", False),
    ("Не плутати: пункт 12 (карантин COVID) ПРОДОВЖУЄ строки і містить перелік "
     "статей 257, 258, 362, 559, 681, 728, 786, 1293. Пункт 19 (воєнний стан) "
     "ЗУПИНЯЄ перебіг і жодного переліку статей не містить. Це різні механізми, "
     "і задачі спеціально перевіряють, чи не сплутано їх.", False),
    ("", False),
    ("Норми, на які спираються задачі", True),
    ("", False),
    ("ст. 257 ЦК — загальна позовна давність три роки", False),
    ("ст. 258 ч. 2 п. 1 ЦК — один рік для вимог про стягнення неустойки "
     "(штрафу, пені)", False),
    ("ст. 259 ч. 1 ЦК — може бути ЗБІЛЬШЕНА за домовленістю, письмово", False),
    ("ст. 259 ч. 2 ЦК — НЕ МОЖЕ бути скорочена за домовленістю", False),
    ("ст. 261 ч. 5 ЦК — за зобов'язаннями з визначеним строком виконання "
     "перебіг починається зі спливом строку виконання", False),
    ("ст. 267 ч. 3, 4 ЦК — застосовується лише за заявою сторони, зробленою до "
     "винесення рішення", False),
    ("ст. 625 ч. 2 ЦК — індекс інфляції плюс три проценти річних", False),
    ("", False),
    ("Про матеріали", True),
    ("", False),
    ("Усі сторони, адреси та коди ЄДРПОУ вигадані. Коди навмисно НЕ проходять "
     "контрольну суму, щоб не збігтися з реальною юридичною особою. Якщо "
     "побачите код, схожий на справжній, це помилка і про неї треба сказати.", False),
]


def load(tasks_root: Path, task_path: str):
    cfg = json.loads((tasks_root / task_path / "task.json").read_text(encoding="utf-8"))
    return cfg


def build(tasks_root: Path, out: Path) -> None:
    out.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Інструкція"
    for i, (line, bold) in enumerate(GUIDE, 1):
        c = ws.cell(row=i, column=1, value=line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if bold:
            c.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 115

    plan = wb.create_sheet("План")
    plan.append(["Етап", "Тема", "Задач", "Критеріїв", "Орієнтовно, хв", "Матеріали справи"])
    for c in plan[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")

    total_min = total_crit = 0
    for no, label, items, minutes in STAGES:
        n = 0
        for tp, sl in items:
            cr = load(tasks_root, tp)["criteria"]
            n += len(cr[sl[0]:sl[1]] if sl else cr)
        total_min += minutes
        total_crit += n
        plan.append([no, label, len(items), n, minutes, f"етап-{no}/"])
    plan.append([])
    plan.append(["РАЗОМ", "", len(STAGES), total_crit, total_min,
                 f"{total_min/60:.1f} год"])
    for col, w in zip("ABCDEF", (7, 52, 8, 11, 15, 22)):
        plan.column_dimensions[col].width = w
    for row in plan.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    dv = DataValidation(type="list", formula1='"OK,ПРАВКА,ПОМИЛКА"', allow_blank=True)

    for no, label, items, minutes in STAGES:
        sh = wb.create_sheet(f"Етап {no}")
        sh.append(HDR)
        for c in sh[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
            c.alignment = Alignment(wrap_text=True, vertical="center")
        stage_dir = out / f"етап-{no}"
        for tp, sl in items:
            cfg = load(tasks_root, tp)
            crits = cfg["criteria"]
            crits = crits[sl[0]:sl[1]] if sl else crits
            for cr in crits:
                sh.append([no, cfg["title"], cr["id"], cr["title"],
                           cr["match_criteria"], cr.get("source", "expert"), "", ""])
            dest = stage_dir / tp.split("/")[-1]
            dest.mkdir(parents=True, exist_ok=True)
            for f in (tasks_root / tp / "documents").iterdir():
                shutil.copy2(f, dest / f.name)
            (dest / "ЗАВДАННЯ.txt").write_text(
                f"{cfg['title']}\n\n{cfg['instructions']}\n\n"
                f"Файл-результат: {', '.join(cfg.get('deliverables', {}))}\n",
                encoding="utf-8")
        sh.add_data_validation(dv)
        dv.add(f"G2:G{sh.max_row}")
        for col, w in zip("ABCDEFGH", (6, 34, 8, 42, 88, 9, 13, 40)):
            sh.column_dimensions[col].width = w
        for row in sh.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        sh.freeze_panes = "A2"

    path = out / "ВАЛІДАЦІЯ-критеріїв.xlsx"
    wb.save(path)
    print(f"{path}")
    print(f"stages={len(STAGES)} criteria={total_crit} minutes={total_min} "
          f"({total_min/60:.1f} h)")


if __name__ == "__main__":
    build(Path(sys.argv[1]), Path(sys.argv[2]))
