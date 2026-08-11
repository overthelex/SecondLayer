#!/usr/bin/env python3
"""Build the lawyer review pack for the Ukrainian LAB tasks.

Produces one .xlsx per task with every criterion on a row, plus the decision
columns a reviewer fills in. The accept rate computed from the filled sheet is
what gets reported in the PR, so the columns are deliberately narrow: a
criterion is either legally correct as written, wrong, or needs an edit.

Usage:
    uv run --with openpyxl python make_review_pack.py <tasks_root> <out_dir>
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import glob as _glob


def discover(tasks_root: Path):
    """Every UA task, diligence first, then litigation."""
    out = []
    for f in sorted(_glob.glob(str(tasks_root / "**" / "task.json"), recursive=True)):
        cfg = json.loads(Path(f).read_text(encoding="utf-8"))
        if cfg.get("jurisdiction") != "UA":
            continue
        rel = Path(f).parent.relative_to(tasks_root).as_posix()
        out.append((rel, cfg["title"]))
    return sorted(out, key=lambda x: (not x[0].startswith("diligence"), x[0]))

HEADER = [
    "ID", "Що перевіряє критерій (EN)", "Формулювання для судді (EN)",
    "Джерело", "ВЕРДИКТ", "Правильне формулювання / коментар",
]

GUIDE = [
    "Як вичитувати.",
    "",
    "Кожен критерій оцінюється як «так/ні» щодо ОДНОГО файлу-результату. "
    "Задача зараховується, лише якщо пройдені ВСІ критерії, тому жоден критерій "
    "не має бути спірним або оціночним.",
    "",
    "У колонці ВЕРДИКТ оберіть одне з:",
    "  OK       - юридично правильно і однозначно, залишаємо як є",
    "  ПРАВКА   - по суті правильно, але формулювання треба уточнити "
    "(напишіть як саме в останній колонці)",
    "  ПОМИЛКА  - юридично неправильно, критерій треба прибрати або "
    "переписати (поясніть чому)",
    "",
    "Особлива увага до критеріїв із джерелом «oracle»: вони мають перевірятися "
    "механічно за текстом закону, тож будь-яка неоднозначність там - це "
    "помилка.",
    "",
    "Контекст щодо часу (стосується задач про позовну давність): кожна справа "
    "умисно датована ДО 14.05.2025, і застосовувати треба редакцію "
    "законодавства, чинну на дату рішення суду першої інстанції, яка вказана в "
    "інструкції до задачі та на аркуші «Задача». Пункт 19 Прикінцевих та "
    "перехідних положень ЦК (зупинення перебігу позовної давності на час "
    "воєнного стану) був чинним на ті дати і виключений лише 14.05.2025 "
    "Законом № 4434-IX. Тому посилання на нього - не помилка, а суть задачі.",
    "",
    "Не плутати: пункт 12 (карантин COVID) ПРОДОВЖУЄ строки і містить перелік "
    "статей 257, 258, 362, 559, 681, 728, 786, 1293. Пункт 19 (воєнний стан) "
    "ЗУПИНЯЄ перебіг і жодного переліку статей не містить.",
    "",
    "Усі сторони, адреси та коди ЄДРПОУ вигадані. Коди навмисно не проходять "
    "контрольну суму, щоб не збігтися з реальною юридичною особою.",
]


def build(tasks_root: Path, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    total = 0
    for task_id, short in discover(tasks_root):
        cfg = json.loads(
            (tasks_root / task_id / "task.json").read_text(encoding="utf-8")
        )
        wb = Workbook()

        ws = wb.active
        ws.title = "Інструкція"
        for i, line in enumerate(GUIDE, 1):
            ws.cell(row=i, column=1, value=line).alignment = Alignment(wrap_text=True)
        ws.column_dimensions["A"].width = 110
        ws.cell(row=1, column=1).font = Font(bold=True, size=13)

        ws2 = wb.create_sheet("Критерії")
        ws2.append(HEADER)
        for c in ws2[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
            c.alignment = Alignment(wrap_text=True, vertical="center")

        for cr in cfg["criteria"]:
            ws2.append([
                cr["id"], cr["title"], cr["match_criteria"],
                cr.get("source", "expert"), "", "",
            ])
            total += 1

        dv = DataValidation(
            type="list", formula1='"OK,ПРАВКА,ПОМИЛКА"', allow_blank=True
        )
        ws2.add_data_validation(dv)
        dv.add(f"E2:E{ws2.max_row}")

        for col, width in zip("ABCDEF", (9, 46, 92, 10, 14, 46)):
            ws2.column_dimensions[col].width = width
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws3 = wb.create_sheet("Задача")
        ws3.append(["Поле", "Значення"])
        ws3["A1"].font = ws3["B1"].font = Font(bold=True)
        for k in ("title", "work_type", "language", "jurisdiction", "judge_language"):
            ws3.append([k, str(cfg.get(k, ""))])
        ws3.append(["instructions", cfg["instructions"]])
        ws3.append(["deliverables", ", ".join(cfg.get("deliverables", {}))])
        ws3.append(["документи", ", ".join(
            sorted(p.name for p in (tasks_root / task_id / "documents").iterdir())
        )])
        ws3.column_dimensions["A"].width = 18
        ws3.column_dimensions["B"].width = 100
        for row in ws3.iter_rows(min_row=2):
            row[1].alignment = Alignment(wrap_text=True, vertical="top")

        name = task_id.split("/")[-1]
        path = out_dir / f"review-{name}.xlsx"
        wb.save(path)
        print(f"{path}  ({len(cfg['criteria'])} criteria)  [{short}]")
    print(f"\ntotal criteria to review: {total}")


if __name__ == "__main__":
    build(Path(sys.argv[1]), Path(sys.argv[2]))
