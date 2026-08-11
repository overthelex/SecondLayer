#!/usr/bin/env python3
"""Rebuild the advocate's package: statute-review tasks plus the three temporal windows.

The previous package went out of date and, worse, was scoped wrong: it covered 40 of the 231
propositions that need a lawyer, none of them from the statute-review family, and spent three of
its eight stages on diligence tasks whose truth lives entirely inside synthetic documents and
needs no legal judgement at all.

What a reviewer is asked here is one question per row: is this proposition of Ukrainian law
correct? Each row therefore carries the exact words of the provision beside the claim, taken
from the dated edition the task is built on, so nothing has to be looked up. That is the single
biggest saving available — the old package made him find the text himself.

Output: ВАЛІДАЦІЯ-2.xlsx with one sheet per stage, plus per-task folders holding the drafted
instrument and the cited provisions only (not the whole act, which runs to 36k words).
"""

import json
import re
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import ua_review
import ua_statut
from ua_claims import CLAIMS, WINDOW_CLAIMS

HERE = Path(__file__).parent
TASKS = Path("harvey-labs/tasks")
ACTS = json.loads((HERE / "acts.json").read_text(encoding="utf-8"))
ACT2275 = json.loads((HERE / "act_2275.json").read_text(encoding="utf-8"))
CK = json.loads((HERE / "ck_provisions.json").read_text(encoding="utf-8"))

VERDICT = '"OK,ПРАВКА,ПОМИЛКА"'
HEAD = PatternFill("solid", fgColor="1F3A34")
HEADF = Font(color="FFFFFF", bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def article_text(nreg, art):
    """The provision as it stands in the edition the task uses."""
    if nreg == "2275-19":
        for a in ACT2275["articles"]:
            if a["n"] == art:
                return a["text"]
        return ""
    act = ACTS.get(nreg)
    if not act:
        return ""
    for a in act["articles"]:
        if a["n"] == art:
            return a["text"]
    return ""


def clean(t, cap=1500):
    t = re.sub(r"\{[^{}]*\}", "", t or "")
    t = re.sub(r"https?://\S+", "", t)
    t = re.sub(r"[ \t]+", " ", t)
    return t.strip()[:cap]


rows = []   # (task, kind, clause, cite, provision_text, claim, required)

# ---- seven tasks from the generic review generator: fully structured ----
for spec in ua_review.SPECS:
    act = ACTS[spec.nreg]
    tag = f"{act['name']} ({spec.nreg}), ред. {ua_review.ed_human(act['ed_date'])}"
    for df in spec.defects:
        rows.append((spec.slug, "НЕВІДПОВІДНІСТЬ", df.clause, f"{df.cite} — {tag}",
                     clean(article_text(spec.nreg, df.art)),
                     f"Пункт проєкту: «{df.text}»",
                     CLAIMS.get((spec.slug, df.clause),
                                f"[БЕЗ ПЕРЕКЛАДУ] {df.breach}. {df.fix}")))
    for lw in spec.lawful:
        rows.append((spec.slug, "ЗАКОННО (пастка)", lw.clause, f"{lw.cite} — {tag}",
                     clean(article_text(spec.nreg, lw.art)),
                     f"Пункт проєкту: «{lw.text}»",
                     CLAIMS.get((spec.slug, lw.clause),
                                f"[БЕЗ ПЕРЕКЛАДУ] {lw.why}")))
    if spec.omission:
        o = spec.omission
        rows.append((spec.slug, "ВІДСУТНЄ", "-", f"{o.cite} — {tag}",
                     clean(article_text(spec.nreg, o.art)),
                     "У проєкті цього положення немає.",
                     CLAIMS.get((spec.slug, "-"),
                                f"[БЕЗ ПЕРЕКЛАДУ] {o.missing}. {o.add}")))

# ---- the pilot, whose generator predates the structured form ----
PILOT = "ua-statut-tov-compliance-review"
ptag = (f"Про товариства з обмеженою та додатковою відповідальністю (2275-19), "
        f"ред. {ua_statut.ACT and ''}01.01.2026")
ART_RE = re.compile(r"статт[іяї]\s+(\d+)")
for df in ua_statut.DEFECTS + [ua_statut.OMISSION]:
    m = ART_RE.search(df.article)
    art = int(m.group(1)) if m else 0
    kind = "ВІДСУТНЄ" if df is ua_statut.OMISSION else "НЕВІДПОВІДНІСТЬ"
    rows.append((PILOT, kind, df.clause, f"{df.article} — {ptag}",
                 clean(article_text("2275-19", art)),
                 f"Пункт проєкту: «{df.text}»",
                 CLAIMS.get((PILOT, df.clause),
                            f"[БЕЗ ПЕРЕКЛАДУ] {df.breach}. {df.fix}")))
for lw in ua_statut.LAWFUL:
    m = ART_RE.search(lw.article)
    rows.append((PILOT, "ЗАКОННО (пастка)", lw.clause, f"{lw.article} — {ptag}",
                 clean(article_text("2275-19", int(m.group(1)) if m else 0)),
                 f"Пункт проєкту: «{lw.text}»",
                 CLAIMS.get((PILOT, lw.clause), f"[БЕЗ ПЕРЕКЛАДУ] {lw.why}")))

# ---- the three windows: propositions live in the criteria, provisions in the Civil Code ----
WINDOWS = {
    "ua-limitation-window-before-p19": ("10.03.2022", "20220101"),
    "ua-limitation-window-original-p19": ("26.06.2023", "20230610"),
    "ua-limitation-window-after-repeal": ("12.11.2025", "20251005"),
}
for slug, (matter, ed) in WINDOWS.items():
    prov = []
    for n in ("12", "19"):
        t = CK.get(ed, {}).get("transitional", {}).get(n)
        prov.append(clean(t, 900) if t else f"(пункт {n} у цій редакції відсутній)")
    for art in ("254", "257", "261"):
        a = CK.get(ed, {}).get("articles", {}).get(art)
        if a:
            prov.append(clean(a, 500))
    rows.append((slug, "ТЕМПОРАЛЬНЕ", f"справа {matter}",
                 f"ЦК України, редакція станом на {ed[6:]}.{ed[4:6]}.{ed[:4]}",
                 "\n\n".join(prov),
                 "Питання: яка редакція діяла на дату рішення і що з неї випливає.",
                 WINDOW_CLAIMS[slug]))

print(f"propositions: {len(rows)}")
by_task = {}
for r in rows:
    by_task.setdefault(r[0], []).append(r)
for t, rs in sorted(by_task.items(), key=lambda kv: -len(kv[1])):
    print(f"  {t[:46]:46s} {len(rs):3d}")

# ---- stages of roughly an hour: ~25 propositions, never splitting a task ----
stages, cur, n = [], [], 0
for t, rs in sorted(by_task.items(), key=lambda kv: kv[0]):
    if n and n + len(rs) > 28:
        stages.append(cur)
        cur, n = [], 0
    cur.append((t, rs))
    n += len(rs)
if cur:
    stages.append(cur)
print(f"\nstages: {len(stages)}")

out = Path("riabchuk2")
if out.exists():
    shutil.rmtree(out)
out.mkdir()

wb = Workbook()
wb.remove(wb.active)
ws = wb.create_sheet("Інструкція")
for i, line in enumerate([
    ("Що і навіщо", True),
    ("", False),
    ("Ми будуємо набір задач, на яких перевіряють юридичних ШІ-агентів. Відповідь агента "
     "оцінюється за списком критеріїв, і задача зараховується, лише якщо пройдено ВСІ критерії. "
     "Тому одне неправильне юридичне твердження псує цілу задачу.", False),
    ("", False),
    ("Ви перевіряєте наше ТВЕРДЖЕННЯ ПРО ПРАВО, а не роботу агента.", True),
    ("Питання одне: чи правильне те, що ми стверджуємо у стовпці «Наше твердження»?", False),
    ("", False),
    ("OK — юридично правильно, нічого не міняти.", False),
    ("ПРАВКА — по суті правильно, але формулювання треба уточнити; напишіть як.", False),
    ("ПОМИЛКА — юридично неправильно; напишіть чому.", False),
    ("", False),
    ("Текст норми наведено поруч, у стовпці «Норма (текст)», станом на редакцію, "
     "чинну на дату справи. Шукати нічого не треба.", False),
    ("", False),
    ("Три типи рядків:", True),
    ("НЕВІДПОВІДНІСТЬ — ми кажемо, що пункт проєкту суперечить нормі.", False),
    ("ЗАКОННО (пастка) — пункт виглядає агресивно, але Закон його дозволяє. Ми навмисно "
     "перевіряємо, чи не оголосить агент його порушенням. Тут питання: чи справді дозволяє?", False),
    ("ВІДСУТНЄ — обов'язкового положення у проєкті немає, і ми кажемо, що воно потрібне.", False),
    ("ТЕМПОРАЛЬНЕ — твердження про те, яка редакція норми діяла на дату справи.", False),
    ("", False),
    ("Окремо: у задачі ua-limitation-window-after-repeal ми СВІДОМО не стверджуємо, що "
     "стається зі строком, який був зупинений до виключення пункту 19. Якщо у Вас є позиція "
     "щодо цього — напишіть, це найцінніше, що ми можемо отримати.", False),
], 1):
    c = ws.cell(row=i, column=1, value=line[0])
    if line[1]:
        c.font = Font(bold=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 118

COLS = ["Задача", "Тип", "Пункт / дата", "Норма", "Норма (текст)",
        "Що у проєкті", "Наше твердження", "ВЕРДИКТ", "Коментар"]
WIDTHS = [30, 18, 16, 34, 60, 52, 62, 14, 44]

for si, stage in enumerate(stages, 1):
    ws = wb.create_sheet(f"Етап {si}")
    for j, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font, c.alignment = HEAD, HEADF, WRAP
        ws.column_dimensions[c.column_letter].width = w
    r = 2
    for task, rs in stage:
        for row in rs:
            for j, v in enumerate(row, 1):
                ws.cell(row=r, column=j, value=v).alignment = WRAP
            ws.row_dimensions[r].height = 78
            r += 1
        # per-stage materials: the drafted instrument only, not the whole act
        src = next(iter(TASKS.glob(f"*/{task}/documents")), None)
        if src:
            dst = out / f"етап-{si}" / task
            dst.mkdir(parents=True, exist_ok=True)
            for f in src.iterdir():
                if f.name.startswith(("proyekt-", "zapyt-", "vidomosti-", "vytyag-",
                                      "dogovir", "pozovna", "rishennya", "vidzyv",
                                      "hronolohiya")):
                    shutil.copy2(f, dst / f.name)
    dv = DataValidation(type="list", formula1=VERDICT, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"H2:H{r}")
    ws.freeze_panes = "A2"

wb.save(out / "ВАЛІДАЦІЯ-2.xlsx")
files = sum(1 for _ in out.rglob("*") if _.is_file())
size = sum(f.stat().st_size for f in out.rglob("*") if f.is_file())
print(f"\nwritten {out}/  — {files} files, {size/1024/1024:.1f} MB")
for si, stage in enumerate(stages, 1):
    print(f"  етап {si}: {sum(len(rs) for _, rs in stage):3d} propositions  "
          f"({', '.join(t[:34] for t, _ in stage)})")
