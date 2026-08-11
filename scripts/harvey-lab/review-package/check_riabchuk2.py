#!/usr/bin/env python3
"""Is every row usable without looking anything up?

The whole point of the rebuild is that the reviewer never has to find a provision himself. A row
with an empty "Норма (текст)" defeats that, and a row whose claim column reads like a rubric
line rather than a proposition of law wastes his attention. Both are checked here, and a few
rows of each type are printed in full so the wording can be judged rather than counted.
"""

import openpyxl

wb = openpyxl.load_workbook("riabchuk2/ВАЛІДАЦІЯ-2.xlsx")
print("sheets:", wb.sheetnames)

empty_prov = 0
total = 0
by_kind = {}
samples = {}
for name in wb.sheetnames:
    if not name.startswith("Етап"):
        continue
    ws = wb[name]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        task, kind, clause, cite, prov, draft, claim = r[:7]
        total += 1
        by_kind[kind] = by_kind.get(kind, 0) + 1
        if not (prov or "").strip():
            empty_prov += 1
            samples.setdefault("EMPTY PROVISION", []).append(r)
        samples.setdefault(kind, []).append(r)

print(f"\nrows: {total}   rows with no provision text: {empty_prov}")
for k, v in sorted(by_kind.items()):
    print(f"  {k:22s} {v}")

for kind in ("НЕВІДПОВІДНІСТЬ", "ЗАКОННО (пастка)", "ВІДСУТНЄ", "ТЕМПОРАЛЬНЕ",
             "EMPTY PROVISION"):
    rs = samples.get(kind, [])
    if not rs:
        continue
    print("\n" + "=" * 96)
    print(f"{kind}  (showing 1 of {len(rs)})")
    task, k, clause, cite, prov, draft, claim = rs[0][:7]
    print(f"  задача:     {task}")
    print(f"  пункт:      {clause}")
    print(f"  норма:      {cite}")
    print(f"  текст:      {(prov or '')[:220]}")
    print(f"  у проєкті:  {(draft or '')[:220]}")
    print(f"  твердження: {(claim or '')[:300]}")
